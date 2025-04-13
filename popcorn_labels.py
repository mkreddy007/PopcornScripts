import labels
import time
import os.path
from reportlab.graphics import shapes
from openpyxl import load_workbook

DEBUG = False

specs = labels.Specification(216, 280,  # Sheet size in MM
                             2, 5,  # Columns then rows
                             101.6, 50.8,  # label size in MM
                             corner_radius=2,  # Is this needed? I think not
                             top_margin=13,  # Top margin in MM
                             bottom_margin=13,  # Bottom margin in MM
                             column_gap=5,  # Margin between Columns
                             row_gap=0)  # No gap between labels

# Array of Teachers that have a teacher helper so they need two bags of popcorn
TwoTeachers = ["Davis", "Tharp", "Espich", "Najjar", "Chamness", "Brookshire", "Park", "Ellis"]
# Array of Teachers that have two?! teacher helpers so they need three bags of popcorn
# ThreeTeachers = ["Estes"]
ThreeTeachers = []
# Dictionary to help pretty print of grade
GradeDict = {"K": "KG", "1": "1st", "2": "2nd", "3": "3rd", "4": "4th", "5": "5th"}
TeacherDict = {"Berrios": "5",
               "Bogin": "1",
               "Brightwell": "1",
               "Brookshire": "4",
               "Chamness": "3",
               "Corbett": "4",
               "Davis": "K",
               "Dunagan": "1",
               "Ellis": "4",
               "Espich": "2",
               "Falsone": "2",
               "Grant": "3",
               "Griffith": "2",
               "Hemphill": "3",
               "Hernandez": "4",
               "Hurst": "K",
               "Martin": "1",
               "Martinez": "K",
               "Najjar": "5",
               "Park": "5",
               "Platt": "4",
               "Richardson": "2",
               "Schiff": "5",
               "Slawson": "2",
               "Tharp": "3",
               }

# Create a function to draw each label. This will be given the ReportLab drawing
# object to draw on, the dimensions (NB. these will be in points, the unit
# ReportLab uses) of the label, and the name to put on the tag.


def write_name(label, width, height, name):
    # Break up the name into it's corrects parts
    format, teacher, grade, num_students = name.split(",")
    if format == "A":
        # Print Pretty Grade and teacher name in a bigger font
        label.add(shapes.String(width / 2.0, 100, GradeDict[grade] + ": " + teacher, fontName="Helvetica", fontSize=36, textAnchor="middle"))
        # Print number of Students in a normal font
        label.add(shapes.String(width / 2.0, 60, num_students + " Students", fontName="Helvetica", fontSize=24, textAnchor="middle"))
        # Split up Teacher and Total sum based on TwoTeacher Array
        if (teacher in ThreeTeachers):
            total = int(num_students) + 3  # Add two to total
            # Print 3 Teachers plus total
            label.add(shapes.String(width / 2.0, 20, "+ 3 Teachers = " + str(total) + " Bags", fontName="Helvetica", fontSize=24, textAnchor="middle"))
        elif (teacher in TwoTeachers):
            total = int(num_students) + 2  # Add two to total
            # Print 2 Teachers plus total
            label.add(shapes.String(width / 2.0, 20, "+ 2 Teachers = " + str(total) + " Bags", fontName="Helvetica", fontSize=24, textAnchor="middle"))
        else:
            total = int(num_students) + 1  # Add one to total
            # Print 1 Teachers plus total
            label.add(shapes.String(width / 2.0, 20, "+ 1 Teacher = " + str(total) + " Bags", fontName="Helvetica", fontSize=24, textAnchor="middle"))
    elif format == "B":
        label.add(shapes.String(width / 2.0, 100, teacher, fontName="Helvetica", fontSize=36, textAnchor="middle"))
        label.add(shapes.String(width / 2.0, 60, grade, fontName="Helvetica", fontSize=24, textAnchor="middle"))
        label.add(shapes.String(width / 2.0, 20, num_students, fontName="Helvetica", fontSize=36, textAnchor="middle"))
    elif format == "C":
        label.add(shapes.String(width / 2.0, 100, teacher, fontName="Helvetica", fontSize=36, textAnchor="middle"))
        label.add(shapes.String(width / 2.0, 60, grade, fontName="Helvetica", fontSize=16, textAnchor="middle"))
        label.add(shapes.String(width / 2.0, 20, num_students, fontName="Helvetica", fontSize=36, textAnchor="middle"))


# Create the sheet.
sheet = labels.Sheet(specs, write_name, border=False)

for file in os.listdir("../Downloads/"):
    if file.endswith(".xlsx"):
        print(os.path.join("../Downloads/", file))
        wb = load_workbook(os.path.join("../Downloads/", file))  # load the excel sheet

# wb = load_workbook('../Downloads/a.xlsx') #load the excel sheet
ws = wb.active

# Create a dictionary to store combined orders for each teacher
teacher_orders = {}

# First pass: accumulate all orders
for row in ws.iter_rows(min_row=3, max_col=10, values_only=True):  # Now scan the spreadsheet
    if row[2] in ["SE-POPCORN", "SE-POPCORN-SPRING-ONLY"]:  # Check for both popcorn types
        if row[0] is not None and row[0] != "Unknown":  # Only look at properly specified teachers
            teacher = row[0]  # Only need the last name
            if teacher not in teacher_orders:
                teacher_orders[teacher] = 0
            teacher_orders[teacher] += row[6]  # Add to the running total for this teacher
        else:
            print("ERROR: " + str(row[6]) + " Kids don't have their teacher specified properly")

# Second pass: create labels using combined totals
for iter_grade in ("K", "1", "2", "3", "4", "5"):  # Iterate through grades in order
    for teacher, num_students in teacher_orders.items():
        grade = TeacherDict[teacher]  # Grab the teacher from the dict
        name = "A," + teacher + ',' + str(grade) + ',' + str(num_students)
        if (grade == iter_grade):  # Only add the correct grade
            if DEBUG:
                print(name)
            sheet.add_label(name.strip())  # Add the label for each class

sheet.add_label("B,CDC Staff,Portable Bldg,8 Bags")
sheet.add_label("B,,,")
sheet.add_label("B,,,")
sheet.add_label("B,Cafeteria,1/2 to 3/4 plastic bag,of loose popcorn")
sheet.add_label("B,Specials,(To Office Front Desk),10 Bags")

sheet.add_label("B,Bus Drivers,Deliver to Library,8 Bags")
sheet.add_label("B,Bus Drivers,Deliver to Library,8 Bags")
sheet.add_label("C,Custodial Office,Inside Kinder/1st Workroom,8 Bags")
sheet.add_label("C,Custodial Office,Inside Kinder/1st Workroom,8 Bags")
sheet.add_label("B,Learning Lab,(To Office Front Desk),22 Bags")
sheet.add_label("B,Learning Lab,(To Office Front Desk),22 Bags")
sheet.add_label("B,Library,Deliver to the Library,8 Bags")
sheet.add_label("B,Library,Deliver to the Library,8 Bags")
sheet.add_label("B,Office,(To Office Front Desk),22 Bags")
sheet.add_label("B,Office,(To Office Front Desk),22 Bags")

timestr = time.strftime("%m_%d_%y")
pdfFileName = "popcorn_" + timestr + ".pdf"
# Save the file and we are done.
sheet.save(pdfFileName)  # Save the label sheets as popcorn.pdf
# Print the number of labels and pages
print("{0:d} label(s) output on {1:d} page(s).".format(sheet.label_count, sheet.page_count))
